import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl

class ExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Viewer & Data Entry Form")
        self.root.geometry("900x500")
        
        self.dark_mode = True  # Default mode is dark
        self.file_path = None
        
        # UI Components
        self.create_widgets()
    
    def create_widgets(self):
        # Sidebar Frame
        self.sidebar = tk.Frame(self.root, bg="#3c3f41", width=250)
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)
        
        tk.Label(self.sidebar, text="Insert Row", fg="white", bg="#3c3f41", font=("Arial", 12)).pack(pady=10)
        
        # Entry Fields
        self.entries = {}
        labels = ["Name", "Age"]
        for label in labels:
            tk.Label(self.sidebar, text=label, fg="white", bg="#3c3f41").pack(pady=2)
            entry = tk.Entry(self.sidebar, bg="#2b2b2b", fg="white", insertbackground="white")
            entry.pack(pady=2, fill=tk.X, padx=10)
            self.entries[label] = entry
        
        # Subscription Dropdown
        tk.Label(self.sidebar, text="Subscription", fg="white", bg="#3c3f41").pack(pady=2)
        self.subscription_var = tk.StringVar()
        self.subscription_dropdown = ttk.Combobox(self.sidebar, textvariable=self.subscription_var, values=["Subscribed", "Unsubscribed"])
        self.subscription_dropdown.pack(pady=2, fill=tk.X, padx=10)
        self.subscription_dropdown.current(0)  # Default to Subscribed
        
        # Employment Checkbox
        self.employed_var = tk.BooleanVar()
        self.chk_employed = tk.Checkbutton(self.sidebar, text="Employed", variable=self.employed_var, bg="#3c3f41", fg="white", selectcolor="#3c3f41", activebackground="#3c3f41")
        self.chk_employed.pack(pady=5)
        
        # Insert Button
        self.btn_add = tk.Button(self.sidebar, text="Insert", command=self.add_entry, bg="#6a8759", fg="white")
        self.btn_add.pack(pady=10, padx=10, fill=tk.X)
        
        # File selection button
        self.btn_load = tk.Button(self.sidebar, text="Load Excel File", command=self.load_file, bg="#6a8759", fg="white")
        self.btn_load.pack(pady=10, padx=10, fill=tk.X)
        
        # Modern Toggle Button
        self.toggle_var = tk.BooleanVar(value=True)
        self.btn_toggle_mode = ttk.Checkbutton(self.sidebar, text="Dark Mode", variable=self.toggle_var, command=self.toggle_mode, style="Toggle.TCheckbutton")
        self.btn_toggle_mode.pack(pady=10, padx=10, fill=tk.X)
        
        # Treeview for displaying data
        self.tree_frame = tk.Frame(self.root, bg="#2b2b2b")
        self.tree_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        
        columns = ["Name", "Age", "Subscription", "Employment"]
        self.tree = ttk.Treeview(self.tree_frame, columns=columns, show="headings", style="mystyle.Treeview")
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center", width=150)
        
        self.tree.pack(expand=True, fill=tk.BOTH)
        
        # Style
        self.set_style()
    
    def set_style(self):
        style = ttk.Style()
        bg_color = "#2b2b2b" if self.dark_mode else "#ffffff"
        fg_color = "white" if self.dark_mode else "black"
        
        self.root.configure(bg=bg_color)
        self.sidebar.configure(bg=bg_color)
        self.tree_frame.configure(bg=bg_color)
        
        style.configure("mystyle.Treeview", background=bg_color, foreground=fg_color, rowheight=25, fieldbackground=bg_color)
        style.map("mystyle.Treeview", background=[("selected", "#6a8759")])
        
        # Modern Toggle Button Style
        style.configure("Toggle.TCheckbutton", background=bg_color, foreground=fg_color, font=("Arial", 10))
    
    def toggle_mode(self):
        self.dark_mode = not self.dark_mode
        self.set_style()
    
    def load_file(self):
        file_path = filedialog.askopenfilename(filetypes=[["Excel files", "*.xlsx;*.xls"]])
        if not file_path:
            return
        
        self.file_path = file_path
        self.display_data()
    
    def display_data(self):
        if not self.file_path:
            messagebox.showerror("Error", "No file selected!")
            return
        
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active
        
        # Clear previous treeview rows
        self.tree.delete(*self.tree.get_children())
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            self.tree.insert("", tk.END, values=row)
        
        wb.close()
    
    def add_entry(self):
        if not self.file_path:
            messagebox.showerror("Error", "No file selected!")
            return
        
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active
        
        values = [self.entries[label].get() for label in self.entries]
        values.append(self.subscription_var.get())
        values.append("Employed" if self.employed_var.get() else "Unemployed")
        
        sheet.append(values)
        
        wb.save(self.file_path)
        wb.close()
        
        messagebox.showinfo("Success", "Entry added successfully!")
        self.display_data()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelApp(root)
    root.mainloop()
